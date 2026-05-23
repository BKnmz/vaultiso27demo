"""
Parses ISO27001_Audit_Checklist_demo.xlsx into ChromaDB.
Run once; skips rebuild if Excel file hasn't changed.
"""

import os
import sys
import hashlib
import json
from pathlib import Path

import yaml
import openpyxl
import chromadb
from sentence_transformers import SentenceTransformer

# Windows console encoding fix
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def load_config(path="config.yaml"):
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def file_hash(path):
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def clean(val):
    """Strip whitespace and return empty string for None."""
    if val is None:
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def parse_general_clauses(ws):
    """
    Returns list of dicts: {control_id, control_name, description, audit_questions}
    Groups multiple question rows per sub-clause into one document.

    Normalizations applied to match pipeline clause IDs:
      - 7.5.1 → 7.5  (Excel has sub-clause; pipeline/skill uses parent)
      - 10.1 ↔ 10.2   (Excel uses ISO 27001:2022 order; codebase uses 2013 order)
    """
    _ID_NORMALIZE = {
        "7.5.1": "7.5",
        "10.1":  "10.2",   # Excel 10.1=Continual Improvement → code 10.2
        "10.2":  "10.1",   # Excel 10.2=Nonconformity         → code 10.1
    }

    docs = []
    current_clause_id = ""
    current_clause_name = ""
    current_questions = []

    def flush():
        if current_clause_id and current_questions:
            docs.append({
                "control_id": current_clause_id,
                "control_name": current_clause_name,
                "description": "",
                "audit_questions": "\n".join(current_questions),
                "sheet": "General_Clauses",
            })

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        col1 = clean(row[1])  # sub-clause id + name
        col2 = clean(row[2])  # question text

        if not col1 and not col2:
            continue

        if col1:
            flush()
            current_questions = []
            # Parse "4.1 - Understanding organization..." → id="4.1", name="..."
            # Use only the first line if the cell is multi-line (e.g. 7.5.1/7.5.2/7.5.3)
            first_line = col1.splitlines()[0]
            if " - " in first_line:
                parts = first_line.split(" - ", 1)
                current_clause_id = _ID_NORMALIZE.get(parts[0].strip(), parts[0].strip())
                current_clause_name = parts[1].strip()
            else:
                current_clause_id = _ID_NORMALIZE.get(first_line, first_line)
                current_clause_name = first_line

        if col2:
            current_questions.append(col2)

    flush()
    return docs


def parse_statement_of_applicability(ws):
    """
    Columns (0-indexed): 1=control_id, 2=control_name, 3=applicable,
                          4=description, 5=justification, 6=status
    Skips header rows and section group rows (no description).
    """
    docs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue  # title + header rows
        control_id = clean(row[1])
        control_name = clean(row[2])
        description = clean(row[4])
        justification = clean(row[5])

        # Skip section headers (e.g. "A5", "A6") and empty rows
        if not control_id or not description:
            continue
        # Skip rows that are section group labels (no dot in id like "A5")
        if "." not in control_id:
            continue

        docs.append({
            "control_id": control_id,
            "control_name": control_name,
            "description": description,
            "audit_questions": justification,
            "sheet": "Statement_of_Applicability",
        })
    return docs


def parse_annex_sheet(ws, sheet_name):
    """
    Columns: 0=control_no, 1=control_name, 2=description, 3=audit_questions
    Header is row 1 (index 0=group header, 1=column names).
    """
    docs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue  # skip group header + column header rows
        control_id = clean(row[0])
        control_name = clean(row[1])
        description = clean(row[2])
        audit_questions = clean(row[3])

        if not control_id or not control_name:
            continue

        docs.append({
            "control_id": f"A.{control_id}" if not control_id.startswith("A") else control_id,
            "control_name": control_name,
            "description": description,
            "audit_questions": audit_questions,
            "sheet": sheet_name,
        })
    return docs


def build_document_text(doc):
    """Builds the text string to embed for a document dict."""
    parts = [f"Control: {doc['control_id']} - {doc['control_name']}"]
    if doc["description"]:
        parts.append(f"Requirement: {doc['description']}")
    if doc["audit_questions"]:
        parts.append(f"Audit Questions: {doc['audit_questions']}")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    cfg = load_config()
    excel_path = Path(cfg["rag"]["source_excel"])
    chroma_path = Path(cfg["rag"]["chroma_db_path"])
    collection_name = cfg["rag"]["collection_name"]
    hash_file = chroma_path / "source.hash"

    print(f"Source: {excel_path}")

    # If source Excel is absent, leave any existing index in place and exit cleanly.
    # The pipeline has a built-in fallback for empty/missing RAG results.
    if not excel_path.exists():
        print(
            "ISO 27001 audit checklist not found — skipping RAG index build.\n"
            "The tool will still run using skill templates only (documents will be\n"
            "generated without ISO reference context). To enable full RAG support,\n"
            "place a compatible audit checklist at:\n"
            f"  {excel_path.resolve()}"
        )
        return

    # Check if rebuild is needed
    current_hash = file_hash(excel_path)
    if chroma_path.exists() and hash_file.exists():
        stored_hash = hash_file.read_text().strip()
        if stored_hash == current_hash:
            print("ChromaDB index is up to date. Skipping rebuild.")
            print(f"  Run with --force to rebuild anyway.")
            return

    print("Building ChromaDB index...")
    chroma_path.mkdir(parents=True, exist_ok=True)

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Parse all sheets
    all_docs = []
    all_docs.extend(parse_general_clauses(wb["General_Clauses"]))
    all_docs.extend(parse_statement_of_applicability(wb["Statement_of_Applicability"]))
    all_docs.extend(parse_annex_sheet(wb["A.5_Operational"], "A.5_Operational"))
    all_docs.extend(parse_annex_sheet(wb["A.6_People"], "A.6_People"))
    all_docs.extend(parse_annex_sheet(wb["A.7_Physical"], "A.7_Physical"))
    all_docs.extend(parse_annex_sheet(wb["A._Technical"], "A._Technical"))
    wb.close()

    print(f"  Parsed {len(all_docs)} documents from Excel.")

    # Deduplicate: prefer Annex sheets over SoA for same control_id
    seen = {}
    for doc in all_docs:
        cid = doc["control_id"]
        # Annex sheets take priority (they have richer audit questions)
        if cid not in seen or doc["sheet"] not in ("Statement_of_Applicability", "General_Clauses"):
            seen[cid] = doc
    unique_docs = list(seen.values())
    print(f"  After dedup: {len(unique_docs)} unique controls/clauses.")

    # Embed
    print("  Loading embedding model (all-MiniLM-L6-v2)...")
    model = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2")

    texts = [build_document_text(d) for d in unique_docs]
    print(f"  Embedding {len(texts)} documents...")
    embeddings = model.encode(texts, show_progress_bar=True, batch_size=32)

    # Store in ChromaDB
    client = chromadb.PersistentClient(path=str(chroma_path))

    # Drop and recreate collection for clean rebuild
    try:
        client.delete_collection(collection_name)
    except Exception:
        pass
    collection = client.create_collection(
        name=collection_name,
        metadata={"hnsw:space": "cosine"},
    )

    ids = [f"doc_{i}" for i in range(len(unique_docs))]
    metadatas = [
        {
            "control_id": d["control_id"],
            "control_name": d["control_name"],
            "sheet": d["sheet"],
        }
        for d in unique_docs
    ]

    # ChromaDB batch limit is 5461 — chunk if needed
    batch_size = 500
    for start in range(0, len(unique_docs), batch_size):
        end = min(start + batch_size, len(unique_docs))
        collection.add(
            ids=ids[start:end],
            embeddings=embeddings[start:end].tolist(),
            documents=texts[start:end],
            metadatas=metadatas[start:end],
        )
    print(f"  Stored {len(unique_docs)} documents in ChromaDB collection '{collection_name}'.")

    # Save hash
    hash_file.write_text(current_hash)
    print("Done. Index saved.")

    # Quick smoke test
    print("\nSmoke test — querying 'access control policy':")
    results = collection.query(
        query_embeddings=model.encode(["access control policy"]).tolist(),
        n_results=3,
    )
    for meta, doc in zip(results["metadatas"][0], results["documents"][0]):
        print(f"  [{meta['control_id']}] {meta['control_name']}")


if __name__ == "__main__":
    import sys
    force = "--force" in sys.argv
    if force:
        cfg = load_config()
        hash_file = Path(cfg["rag"]["chroma_db_path"]) / "source.hash"
        if hash_file.exists():
            hash_file.unlink()
        print("Force flag set — rebuilding index.")
    main()
