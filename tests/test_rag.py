"""
Tests for rag_setup.py — Excel parsing and ChromaDB indexing.
Adapted for the demo repo: ISO27001_Audit_Checklist_demo.xlsx (21 Annex A controls).
These tests run without Ollama.
"""

import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from rag_setup import (
    parse_general_clauses,
    parse_statement_of_applicability,
    parse_annex_sheet,
    build_document_text,
    clean,
)
import openpyxl


EXCEL_PATH = Path(__file__).parent.parent / "rag" / "ISO27001_Audit_Checklist_demo.xlsx"


class TestClean(unittest.TestCase):
    def test_none_returns_empty(self):
        self.assertEqual(clean(None), "")

    def test_strips_whitespace(self):
        self.assertEqual(clean("  hello  "), "hello")

    def test_converts_non_string(self):
        self.assertEqual(clean(5.1), "5.1")


class TestGeneralClausesParsing(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        cls.docs = parse_general_clauses(wb["General_Clauses"])
        wb.close()

    def test_returns_list(self):
        self.assertIsInstance(self.docs, list)

    def test_has_documents(self):
        self.assertGreater(len(self.docs), 0)

    def test_clause_41_present(self):
        ids = [d["control_id"] for d in self.docs]
        self.assertIn("4.1", ids)

    def test_clause_43_present(self):
        ids = [d["control_id"] for d in self.docs]
        self.assertIn("4.3", ids)

    def test_required_fields(self):
        for doc in self.docs:
            self.assertIn("control_id", doc)
            self.assertIn("control_name", doc)
            self.assertIn("audit_questions", doc)
            self.assertIn("sheet", doc)

    def test_audit_questions_non_empty(self):
        for doc in self.docs:
            self.assertTrue(
                len(doc["audit_questions"]) > 0,
                f"Empty audit_questions for {doc['control_id']}"
            )

    def test_sheet_label_correct(self):
        for doc in self.docs:
            self.assertEqual(doc["sheet"], "General_Clauses")


class TestSoAParsing(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        cls.docs = parse_statement_of_applicability(wb["Statement_of_Applicability"])
        wb.close()

    def test_returns_list(self):
        self.assertIsInstance(self.docs, list)

    def test_has_annex_a_controls(self):
        # Demo SoA has 22 controls (subset of 93)
        self.assertGreater(len(self.docs), 5)

    def test_control_ids_have_dots(self):
        for doc in self.docs:
            self.assertIn(".", doc["control_id"], f"No dot in {doc['control_id']}")

    def test_a51_present(self):
        ids = [d["control_id"] for d in self.docs]
        self.assertIn("A.5.1", ids)

    def test_no_section_headers(self):
        ids = [d["control_id"] for d in self.docs]
        self.assertNotIn("A5", ids)
        self.assertNotIn("A6", ids)


class TestAnnexSheetParsing(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        cls.docs_a5 = parse_annex_sheet(wb["A.5_Operational"], "A.5_Operational")
        cls.docs_a6 = parse_annex_sheet(wb["A.6_People"], "A.6_People")
        cls.docs_a7 = parse_annex_sheet(wb["A.7_Physical"], "A.7_Physical")
        cls.docs_tech = parse_annex_sheet(wb["A._Technical"], "A._Technical")
        wb.close()

    def test_a5_has_controls(self):
        self.assertGreater(len(self.docs_a5), 0)

    def test_control_ids_prefixed(self):
        for doc in self.docs_a5:
            self.assertTrue(
                doc["control_id"].startswith("A."),
                f"Missing A. prefix: {doc['control_id']}"
            )

    def test_descriptions_non_empty(self):
        non_empty = [d for d in self.docs_a5 if d["description"]]
        self.assertGreater(len(non_empty), 0)

    def test_sheet_label(self):
        for doc in self.docs_a5:
            self.assertEqual(doc["sheet"], "A.5_Operational")


class TestBuildDocumentText(unittest.TestCase):
    def test_includes_control_id(self):
        doc = {"control_id": "5.1", "control_name": "Test", "description": "Desc", "audit_questions": "Q1"}
        text = build_document_text(doc)
        self.assertIn("5.1", text)
        self.assertIn("Test", text)

    def test_empty_description_skipped(self):
        doc = {"control_id": "5.1", "control_name": "Test", "description": "", "audit_questions": "Q1"}
        text = build_document_text(doc)
        self.assertNotIn("Requirement:", text)

    def test_audit_questions_included(self):
        doc = {"control_id": "5.1", "control_name": "Test", "description": "D", "audit_questions": "Do X?"}
        text = build_document_text(doc)
        self.assertIn("Do X?", text)


class TestChromaDBIndex(unittest.TestCase):
    """Tests against the live ChromaDB index built by rag_setup.py."""

    @classmethod
    def setUpClass(cls):
        import chromadb
        from sentence_transformers import SentenceTransformer
        import yaml

        cfg = yaml.safe_load((Path(__file__).parent.parent / "config.yaml").read_text())
        chroma_path = Path(__file__).parent.parent / cfg["rag"]["chroma_db_path"]

        if not chroma_path.exists():
            cls.collection = None
            cls.model = None
            return

        client = chromadb.PersistentClient(path=str(chroma_path))
        cls.collection = client.get_collection(cfg["rag"]["collection_name"])
        cls.model = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2")

    def setUp(self):
        if self.collection is None:
            self.skipTest("ChromaDB index not built — run rag_setup.py first")

    def test_collection_not_empty(self):
        count = self.collection.count()
        # Demo has ~50+ entries (general clauses + 22 SoA + 21 Annex A controls)
        self.assertGreater(count, 20)

    def test_exact_match_clause_43(self):
        r = self.collection.get(where={"control_id": {"$eq": "4.3"}})
        self.assertGreater(len(r["ids"]), 0, "Clause 4.3 not found in index")

    def test_exact_match_a51(self):
        r = self.collection.get(where={"control_id": {"$eq": "A.5.1"}})
        self.assertGreater(len(r["ids"]), 0, "A.5.1 not found in index")

    def test_semantic_query_policy(self):
        emb = self.model.encode(["information security policy"]).tolist()
        r = self.collection.query(query_embeddings=emb, n_results=3)
        ids = [m["control_id"] for m in r["metadatas"][0]]
        # Should return policy-related entries present in demo
        policy_controls = {"A.5.1", "5.2", "5.1", "A.5.2"}
        self.assertTrue(
            any(cid in policy_controls for cid in ids),
            f"Expected policy results, got: {ids}"
        )

    def test_semantic_query_risk_assessment(self):
        emb = self.model.encode(["risk assessment methodology likelihood impact"]).tolist()
        r = self.collection.query(query_embeddings=emb, n_results=3)
        ids = [m["control_id"] for m in r["metadatas"][0]]
        risk_controls = {"6.1", "6.1.2", "6.1.3", "8.2"}
        self.assertTrue(
            any(cid in risk_controls for cid in ids),
            f"Expected risk assessment results, got: {ids}"
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
